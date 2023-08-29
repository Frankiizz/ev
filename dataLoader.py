import openpyxl


def load_excel(filename):
    # Load the Excel file
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.active

    # Get the maximum number of columns in the sheet
    max_column = sheet.max_column

    # Initialize lists to store columns
    columns = [[] for _ in range(max_column)]

    # Iterate through each row and store values in respective columns
    for row in sheet.iter_rows(values_only=True):
        for col_num, value in enumerate(row):
            columns[col_num].append(value)

    return columns

def output_excel():
    pass
